from validation import epochVal_metrics_test
from options import args_parser
import os
import sys
import logging
import random
import numpy as np
import matplotlib.pyplot as plt
import copy
from FedAvg import FedAvg,AvgW,AddW
import torch
from torchvision import transforms
import torch.backends.cudnn as cudnn
from networks.convnext.convnextv2 import ConvNeXtV2, convnextv2_huge,convnextv2_base,convnextv2_large,convnextv2_atto,convnextv2_tiny,convnextv2_nano,convnext_pico
from networks.convnext.convnext import convnext_tiny
import datetime
import openpyxl
from torch.utils.data.sampler import Sampler
from confuse_matrix import get_confuse_matrix, kd_loss,update_pseudo_labels
from sklearn.preprocessing import LabelBinarizer
from test import final_test

from imblearn.over_sampling import SMOTE
from PIL import Image
import pandas as pd

from dataloaders import dataset
from local_supervised import SupervisedLocalUpdate
from local_unsupervised import UnsupervisedLocalUpdate
from torch.utils.data import DataLoader,WeightedRandomSampler,Dataset
from networks.models import DenseNet121

args = args_parser()
class WeightedShuffledSampler(Sampler):
    def __init__(self, weights, num_samples):
        self.weights = weights
        self.num_samples = num_samples

    def __iter__(self):
        return iter(torch.multinomial(torch.tensor(self.weights), self.num_samples, replacement=True))

    def __len__(self):
        return self.num_samples
class DatasetSplit0(Dataset):
    def __init__(self, dataset, idxs):
        self.dataset = dataset
        self.idxs = list(idxs)

    def __len__(self):
        return len(self.idxs)

    def __getitem__(self, item):
        items, index, image, label = self.dataset[self.idxs[item]]
        return items, index, image, label


class DatasetSplit(Dataset):
    def __init__(self, dataset, idxs):
        self.dataset = dataset
        self.idxs = list(idxs)

    def __len__(self):
        return len(self.idxs)

    def __getitem__(self, item):
        items, index, weak_aug, strong_aug, label = self.dataset[self.idxs[item]]
        return items, index, weak_aug, strong_aug, label

def split(dataset, num_users):
    num_items = int(len(dataset) / num_users)
    dict_users, all_idxs = {}, [i for i in range(len(dataset))]
    print(len(dataset))
    for i in range(num_users):
        dict_users[i] = set(np.random.choice(all_idxs, num_items, replace=False))
        all_idxs = list(set(all_idxs) - dict_users[i])
    return dict_users
def overSampling_smote(train_dataset,client):
    labels_arr = []
    images_arr= []
    for idx in dict_users[client]:
        image = train_dataset.dataset.getImage(idx)
        images_arr.append(np.array(image))  # Convert PIL Image to numpy array
        labels_arr.append(train_dataset.dataset.getLabel(idx))
    
    # Convert lists to numpy arrays
    images_arr = np.array(images_arr)
    labels_arr = np.array(labels_arr)
    smote = SMOTE(k_neighbors=4)
    #float_list = [int(num) for num in dict_users[0]]
    
    
    n_samples, width, height, channels = images_arr.shape
    images_reshaped = images_arr.reshape((n_samples, -1))

    # Resample using SMOTE
    smote = SMOTE(k_neighbors=6)
    X_resampled, y_resampled = smote.fit_resample(images_reshaped, labels_arr)

    # Reshape resampled images back to original shape
    X_resampled = X_resampled.reshape((-1, width, height, channels))
    # Create a directory to save the resampled images
    class_names = args.class_names
    label_binarizer = LabelBinarizer()
    label_binarizer.fit(class_names)

    # Create a directory to save the resampled images
    output_dir = f"../data/{args.dataset}/dict{client}_images"
    os.makedirs(output_dir, exist_ok=True)

    # Save images and record their names and one-hot encoded labels
    image_data = []
    for i, img in enumerate(X_resampled):
        img_name = f"dict{client}_image_{i}.jpg"
        img_path = os.path.join(output_dir, img_name)
        # Save the image in JPEG format
        Image.fromarray(img.astype(np.uint8)).save(img_path, "JPEG")
        # Get the one-hot encoded label
        one_hot_label = label_binarizer.transform([class_names[y_resampled[i]]])[0]
        # Record the image name and one-hot encoded label
        image_data.append([img_name] + one_hot_label.tolist())

    # Create a DataFrame and save to Excel
    columns = ["ImageID"] + class_names
    df = pd.DataFrame(image_data, columns=columns)
    df.to_csv(f"../data/{args.dataset}/dict{client}_train.csv", index=False)

    new_indices = range(0,  len(X_resampled))
    dict_users[client] = list(new_indices)
    
    train_dataset_new=makeDataset(output_dir , f"../data/{args.dataset}/dict{client}_train.csv")
    train_dataset_new, _ = torch.utils.data.random_split(train_dataset_new, [1, 0])
    
    print("Images and labels have been saved successfully.")
    return train_dataset_new

def makeDataset(root_dir,csv_file):
        normalize = transforms.Normalize([0.485, 0.456, 0.406], [0.229, 0.224, 0.225])

        train_dataset = dataset.CheXpertDataset(
            root_dir=root_dir,
            csv_file=csv_file,
            transform=dataset.TransformTwice(
                transforms.Compose(
                    [
                        transforms.Resize((224, 224)),
                        transforms.RandomAffine(degrees=10, translate=(0.02, 0.02)),
                        transforms.RandomHorizontalFlip(),
                        transforms.ToTensor(),
                        normalize,
                    ]   
                )
            ),
        )
        return train_dataset
def GroupIndecies(index):
    if index<7:
        x=0
        y=index+1
    elif index <13:
        x=1
        y=index-5
    elif index <18:
        x=2
        y=index-10
    elif index <22:
        x=3
        y=index-14
    elif index <25:
        x=4
        y=index-17
    elif index <27:
        x=5
        y=index-19
    elif index <28:
        x=6
        y=index-20
    return x,y
def test(epoch, w_glob):
    #checkpoint_path = save_mode_path

    #checkpoint = torch.load(checkpoint_path)
    
    #net = convnextv2_tiny( args.num_classes, mode=args.label_uncertainty)

    if args.network=="convnext":
        net= convnext_tiny(args.num_classes,pretrained=True, in_22k=True)
    else:
        net = DenseNet121(out_size=args.num_classes, mode=args.label_uncertainty, drop_rate=args.drop_rate)
    model = net.cuda()
    if len(args.gpu.split(",")) > 1:
        model = torch.nn.DataParallel(net)
    model.load_state_dict(w_glob)
    #model.load_state_dict(checkpoint["state_dict"])
    normalize = transforms.Normalize([0.485, 0.456, 0.406], [0.229, 0.224, 0.225])
    test_dataset = dataset.CheXpertDataset(
        root_dir=args.root_path,
        csv_file=args.csv_file_test,
        transform=transforms.Compose(
            [
                transforms.Resize((224, 224)),
                transforms.ToTensor(),
                normalize,
            ]
        ),
    )
    test_dataloader = DataLoader(
        dataset=test_dataset,
        batch_size=args.batch_size,
        shuffle=False,
        num_workers=2,
        pin_memory=True,
    )
    AUROCs, accuracy,Accus, Senss, Specs, _, _,_ = epochVal_metrics_test(
        model, test_dataloader, thresh=0.4
    )
    AUROC_avg = np.array(AUROCs).mean()
    Accus_avg = np.array(Accus).mean()
    Senss_avg = np.array(Senss).mean()
    Specs_avg = np.array(Specs).mean()

    return AUROC_avg, accuracy,Accus_avg, Senss_avg, Specs_avg

def saveTraningValidationResults(new_folder_path,accuracy_avg,sup_accuracy_avg,unsup_accuracy_avg,val_accuracy,val_accuracy_thres,used_lp_ratio_avg,loss_avg,sup_loss_avg,unsup_loss_avg):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.cell(row=1, column=1).value = 'Round'
    ws.cell(row=2 ,column=1).value = 'Train Acc'
    ws.cell(row=3, column=1).value = 'Sup Train Acc'
    ws.cell(row=4, column=1).value = 'Unsup Train Acc'
    ws.cell(row=5, column=1).value = 'val Acc'
    ws.cell(row=6, column=1).value = 'val Acc Thr'
    ws.cell(row=7, column=1).value = 'used pl ratio'
    ws.cell(row=9, column=1).value = 'Train avg'
    ws.cell(row=10, column=1).value = 'Sup Train loss'
    ws.cell(row=11, column=1).value = 'Unsup Train loss'

    for i in range(com_round):
        ws.cell(row=1, column=i + 2).value = i
        ws.cell(row=2, column=i + 2).value = "{:.4f}".format(accuracy_avg[i])
        ws.cell(row=3, column=i + 2).value = "{:.4f}".format(sup_accuracy_avg[i])
        ws.cell(row=4, column=i + 2).value = "{:.4f}".format(unsup_accuracy_avg[i])
        ws.cell(row=5, column=i + 2).value = "{:.4f}".format(val_accuracy[i])
        ws.cell(row=6, column=i + 2).value = "{:.4f}".format(val_accuracy_thres[i])
        ws.cell(row=7, column=i + 2).value = "{:.4f}".format(used_lp_ratio_avg[i])
        ws.cell(row=9, column=i + 2).value = "{:.4f}".format(loss_avg[i])
        ws.cell(row=10, column=i + 2).value = "{:.4f}".format(sup_loss_avg[i])
        ws.cell(row=11, column=i + 2).value = "{:.4f}".format(unsup_loss_avg[i])

    wb.save(os.path.join(new_folder_path, 'datasheet.xlsx'))




    plt.plot(accuracy_avg, label='Train Acc')
    plt.plot(sup_accuracy_avg, label='Sup Train Acc')
    plt.plot(unsup_accuracy_avg, label='Unsup Train Acc')
    plt.plot(val_accuracy, label='val Acc')
    plt.plot(val_accuracy_thres, label='val Acc Thr')
    plt.xlabel('Round')
    plt.ylabel('Accuracy')
    plt.legend()
    plt.savefig(os.path.join(new_folder_path, 'accuracy.png'))
    plt.clf()

    plt.plot(loss_avg, label='Train loss')
    plt.plot(sup_loss_avg, label='Sup Train loss')
    plt.plot(unsup_loss_avg, label='Unsup Train loss')
    plt.xlabel('Round')
    plt.ylabel('Loss')
    plt.legend()
    plt.savefig(os.path.join(new_folder_path, 'loss.png'))

AUROCs = []
Accus = []
Senss = []
Specs = []

match args.mode:
    
    case "sup":
        supervised_user_id = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
        unsupervised_user_id = []
    case "ssfl":
        supervised_user_id = [0, 1]
        unsupervised_user_id = [2, 3, 4, 5, 6, 7, 8,9]
    case "unsup":
        supervised_user_id = []
        unsupervised_user_id = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]

flag_create = False
print("done")
useSMOTE=False
dict0_train_dataset=None
dict1_train_dataset=None
if __name__ == "__main__":
    snapshot_folder = os.path.join("..", "model", f"{args.dataset}_{args.mode}")
    os.makedirs(snapshot_folder, exist_ok=True)

    # create logs folder if not exist
    os.makedirs(os.path.join("..", "logs"), exist_ok=True)

    os.environ["CUDA_VISIBLE_DEVICES"] = args.gpu

    results_dir = 'results'
    now = datetime.datetime.now()
    folder_name = now.strftime(f"%Y-%m-%d_%H-%M-%S-{args.com}")
    # Check if the graphs directory exists

    dataset_dir = os.path.join(results_dir, args.dataset)
    if not os.path.exists(results_dir) or not os.path.isdir(results_dir):
        os.makedirs(results_dir)
    if not os.path.exists(dataset_dir) or not os.path.isdir(dataset_dir):
        os.makedirs(dataset_dir)
    new_folder_path = os.path.join(dataset_dir, folder_name)
    if not os.path.exists(new_folder_path):
        os.makedirs(new_folder_path)

    #logging.basicConfig(
    #    filename=os.path.join("..", "logs", f"{args.dataset}_{args.mode}.log"),
    #   level=logging.INFO,
    #    format="[%(asctime)s.%(msecs)03d] %(message)s",
    #    datefmt="%H:%M:%S",
    #)
    logging.basicConfig(
        filename=os.path.join(results_dir,args.dataset,folder_name, f"{args.dataset}_{args.mode}.log"),
        level=logging.INFO,
        format="[%(asctime)s.%(msecs)03d] %(message)s",
        datefmt="%H:%M:%S",
    )
    logging.getLogger().addHandler(logging.StreamHandler(sys.stdout))

    # add a line end of the log file
    logging.info(f"{'='*100}\n")

    logging.info(str(args))

    if args.deterministic:
        cudnn.benchmark = False
        cudnn.deterministic = True
        random.seed(args.seed)
        np.random.seed(args.seed)
        torch.manual_seed(args.seed)
        if torch.cuda.is_available():
            torch.cuda.manual_seed(args.seed)
        else:
            torch.manual_seed(args.seed)
    
    
    train_dataset=makeDataset(args.root_path,args.csv_file_train)
    train_dataset, _ = torch.utils.data.random_split(train_dataset, [0.875, 0.125])
    #train_dataset, _ = torch.utils.data.random_split(train_dataset, [1, 0])
    dict_users = split(train_dataset, args.num_users)
    labels_arr = []
    for idx in dict_users[0]:
        labels_arr.append(train_dataset.dataset.getLabel(idx))
    from collections import Counter

    label_counts = Counter(labels_arr)

    # Display the counts in order of appearance in the list
    label_order = list(dict.fromkeys(labels_arr))
    ordered_counts = {label: label_counts[label] for label in label_order}

    print(ordered_counts)
    labels_arr = []
    for idx in dict_users[1]:
        labels_arr.append(train_dataset.dataset.getLabel(idx))
    from collections import Counter

    label_counts = Counter(labels_arr)

    # Display the counts in order of appearance in the list
    label_order = list(dict.fromkeys(labels_arr))
    ordered_counts = {label: label_counts[label] for label in label_order}

    print(ordered_counts)
###################################################################################################################
###############################################SMOTE###############################################################
###################################################################################################################
    
    if useSMOTE:
        dict0_train_dataset=overSampling_smote(train_dataset,0)
        dict1_train_dataset=overSampling_smote(train_dataset,1)

    #flattened_list = [item for sublist in X_resampled for item in sublist]
    
    #dict_users[0] = set(flattened_list)
    
    # labels_arr = []
    # for idx in dict_users[1]:
    #     labels_arr.append(train_dataset.dataset.getLabel(idx))

    # smote = SMOTE(k_neighbors=4)
    # float_list = [int(num) for num in dict_users[1]]
    # X_resampled, y_resampled = smote.fit_resample(np.array(float_list).reshape(-1, 1), labels_arr)
    # flattened_list = [item for sublist in X_resampled for item in sublist]
    # dict_users[1] = set(flattened_list)

#######################################################################################################

    if args.network=="convnext":
        net_glob= convnext_tiny(args.num_classes,pretrained=True, in_22k=True)
    else:
        net_glob = DenseNet121(out_size=args.num_classes, mode=args.label_uncertainty, drop_rate=args.drop_rate)
    

    #net_glob=net_glob.cuda()
    if len(args.gpu.split(",")) > 1:
        net_glob = torch.nn.DataParallel(net_glob, device_ids=[0, 1])
    net_glob.train()
    w_glob = net_glob.state_dict()
    #w_locals = []
    wgb_locals = []
    trainer_locals = []
    #net_locals = []
    optim_locals = []
    confuse_matrix_mean = []
    confuse_matrix_covariance = []
    unsup_confuse_matrix = []
    unsup_confuse_matrix_index = []
    if torch.cuda.is_available():
        # net_locals.append(copy.deepcopy(net_glob).cuda())
        net_locals = (copy.deepcopy(net_glob).cuda())

        optimizer = torch.optim.Adam(
            net_locals.parameters(),
            lr=args.base_lr,
            betas=(0.9, 0.999),
            weight_decay=5e-4,
        )
    for i in supervised_user_id:
        confuse_matrix_mean.append(torch.zeros((args.num_classes, args.mean_used_features)).cuda())
        confuse_matrix_covariance.append(torch.zeros((args.num_classes, args.covariance_used_features)).cuda())
        wgb_locals.append(copy.deepcopy(w_glob))
        #trainer_locals.append(SupervisedLocalUpdate(args, train_dataset, dict_users[i]))
        



        optim_locals.append(copy.deepcopy(optimizer.state_dict()))

    for i in unsupervised_user_id:
        wgb_locals.append(copy.deepcopy(w_glob))
        #unsup_confuse_matrix.append(torch.zeros((args.num_classes, args.num_classes)).cuda())
        #unsup_confuse_matrix_index.append(-1)
    #        UnsupervisedLocalUpdate(args, train_dataset, dict_users[i])

    trainsup = True
    accuracy_avg = []
    loss_avg = []
    sup_accuracy_avg = []
    sup_loss_avg = []
    unsup_accuracy_avg = []
    unsup_loss_avg = []
    val_accuracy_thres = []
    val_accuracy = []
    used_lp_ratio_avg = []
    w1= copy.deepcopy(net_glob.state_dict())
    w2= copy.deepcopy(net_glob.state_dict())
    w3= copy.deepcopy(net_glob.state_dict())
    w4= copy.deepcopy(net_glob.state_dict())
    w5= copy.deepcopy(net_glob.state_dict())
    w_tem= copy.deepcopy(net_glob.state_dict())
    higher_accuracy=0
    higher_accuracy_round=0
    indecies =[2,8,3,5,4,7,6,9]
    random.shuffle(indecies)
    print(indecies)


    try :
        for com_round in range(args.rounds):
            print("begin")
            loss_locals = []
            Accurcy_locals = []
            sup_loss_locals = []
            sup_Accurcy_locals = []
            unsup_loss_locals = []
            unsup_Accurcy_locals = []
            used_lp_ratio = []
            w_len = 0
            if com_round * args.local_ep < 200:
                for idx in supervised_user_id:
                    train_dataset.dataset.is_labeled=True
                    if useSMOTE:
                        if idx==0:
                            local1 = SupervisedLocalUpdate(args, dict0_train_dataset, dict_users[idx])
                        elif idx==1:
                            local1 = SupervisedLocalUpdate(args, dict1_train_dataset, dict_users[idx])
                    else:
                        local1 = SupervisedLocalUpdate(args, train_dataset, dict_users[idx])

                    if com_round * args.local_ep > 20:
                        local1.base_lr = 3e-4

                    optimizer = optim_locals[idx]

                    net_locals1 = (copy.deepcopy(net_glob).cuda())
                    net_locals1.load_state_dict(w_glob)

                    w, loss,acc, op, confuse_matrix_mean[idx],confuse_matrix_covariance[idx] = local1.train(args, net_locals1, optimizer)
                    wgb_locals[idx]=(copy.deepcopy(w))
                    if (w_len == 0):
                        w_locals = copy.deepcopy(w)
                    else:
                        w_locals = AddW(w_locals, copy.deepcopy(w))
                    w_len += 1
                    

                    
                    
                    
                    #w_locals[idx] = copy.deepcopy(w)
                    optim_locals[idx] = copy.deepcopy(op)

                    loss_locals.append(copy.deepcopy(loss))
                    Accurcy_locals.append(copy.deepcopy(acc))
                    sup_loss_locals.append(copy.deepcopy(loss))
                    sup_Accurcy_locals.append(copy.deepcopy(acc))

            #if com_round>0:
            if com_round * args.local_ep > 20 and len(unsupervised_user_id)>0:
            
                if not flag_create:

                    print("begin unsup")
                    w2= copy.deepcopy(w_glob)
                    w3= copy.deepcopy(w_glob)
                    w4= copy.deepcopy(w_glob)
                    w5= copy.deepcopy(w_glob)
                    net_locals = (copy.deepcopy(net_glob).cuda())

                    optimizer = torch.optim.Adam(
                        net_locals.parameters(),
                        lr=args.base_lr,
                        betas=(0.9, 0.999),
                        weight_decay=5e-4,
                    )
                    ema_models = []
                    ema_models_iternumber = []
                    for i in unsupervised_user_id:
                        #w_locals.append(copy.deepcopy(w_glob))
                       # if torch.cuda.is_available():
                           # net_locals.append(copy.deepcopy(net_glob).cuda())
                        #else:
                        #    net_locals.append(copy.deepcopy(net_glob))
                        
                        if args.network=="convnext":
                            net_ema= convnext_tiny(args.num_classes,pretrained=True, in_22k=True)
                        else:
                            net_ema = DenseNet121(out_size=args.num_classes, mode=args.label_uncertainty, drop_rate=args.drop_rate)
                        #net_ema = convnextv2_tiny(args.num_classes, mode=args.label_uncertainty)
                        # net_ema.head = torch.nn.Sequential(torch.nn.Linear(768, 192),
                        #      torch.nn.ReLU(),
                        #      torch.nn.Linear(192, args.num_classes)
                        #     )
                        if len(args.gpu.split(",")) > 1:
                            net = torch.nn.DataParallel(net_ema.cuda(), device_ids=[0, 1])
                        if torch.cuda.is_available():
                            ema_model = net_ema.cuda()
                        else:
                            ema_model = net_ema
                        for param in ema_model.parameters():
                            param.detach_()
                        ema_models.append(copy.deepcopy(ema_model))
                        ema_models_iternumber.append(0)

                        optim_locals.append(copy.deepcopy(optimizer.state_dict()))
                    flag_create = True
                for idx in unsupervised_user_id:
                    train_dataset.dataset.is_labeled=False

                    #local = trainer_locals[idx]
                    optimizer = optim_locals[idx]
                    
                    net_locals = (copy.deepcopy(net_glob).cuda())
                    if idx==indecies[0] or idx==indecies[1]:
                        net_locals.load_state_dict(w2)
                    if idx==indecies[2] or idx==indecies[3]:
                        net_locals.load_state_dict(w3)
                    if idx==indecies[4] or idx==indecies[5]:
                        net_locals.load_state_dict(w4)
                    if idx==indecies[6] or idx==indecies[7]:
                        net_locals.load_state_dict(w5)
                    #net_locals.load_state_dict(w_glob)
                    local = UnsupervisedLocalUpdate(args, train_dataset, dict_users[idx])
                    # w, loss,acc,ema_models[idx-2], ema_models_iternumber[idx-2], op,unsup_confuse_matrix[idx-2] = local.train(
                    #     args,
                    #     net_locals,ema_models[idx-2], ema_models_iternumber[idx-2],
                    #     optimizer,
                    #     com_round * args.local_ep,
                    #     avg_matrix,unsup_confuse_matrix[idx-2]
                    # )
                    w, loss,acc,ema_models[idx-2], ema_models_iternumber[idx-2], op,usedlp = local.train(
                        args,
                        net_locals,ema_models[idx-2], ema_models_iternumber[idx-2],
                        optimizer,
                        com_round * args.local_ep,
                        avg_matrix_mean,
                        avg_matrix_covariance
                    )
                    wgb_locals[idx]=(copy.deepcopy(w))
                    
                    if (w_len == 0):
                        w_locals = copy.deepcopy(w)
                    else:
                        w_locals = AddW(w_locals, copy.deepcopy(w))
                    w_len += 1
                    
                    
                    #w_locals[idx] = copy.deepcopy(w)
                    optim_locals[idx] = copy.deepcopy(op)
                    loss_locals.append(copy.deepcopy(loss))
                    used_lp_ratio.append(copy.deepcopy(usedlp))
                    Accurcy_locals.append(copy.deepcopy(acc))

                    unsup_loss_locals.append(copy.deepcopy(loss))
                    unsup_Accurcy_locals.append(copy.deepcopy(acc))

            with torch.no_grad():
                avg_matrix_mean = confuse_matrix_mean[0]
                for idx in supervised_user_id[1:]:
                    avg_matrix_mean = avg_matrix_mean + confuse_matrix_mean[idx]
                avg_matrix_mean = avg_matrix_mean / len(supervised_user_id)
            with torch.no_grad():
                avg_matrix_covariance = confuse_matrix_covariance[0]
                for idx in supervised_user_id[1:]:
                    avg_matrix_covariance = avg_matrix_covariance + confuse_matrix_covariance[idx]
                avg_matrix_covariance = avg_matrix_covariance / len(supervised_user_id)
                

            with torch.no_grad():
                if args.aggregator=="gba":
#                     klValues=[]
#                     for i  in range(len(unsupervised_user_id)):
#                         for j in range(i+1,len(unsupervised_user_id)):
#                             klValues.append(kd_loss(unsup_confuse_matrix[i],unsup_confuse_matrix[j]).item())
                            
#                     klValues_tensor = torch.tensor(klValues)

#                     # Sort the indices
#                     sorted_indices = (torch.argsort(klValues_tensor))
#                     i=0
#                     for j in range(len(sorted_indices)):
#                         x,y=GroupIndecies(sorted_indices[j].item())
#                         if x not in unsup_confuse_matrix_index  and y not in unsup_confuse_matrix_index:
#                             unsup_confuse_matrix_index[i]=x
#                             unsup_confuse_matrix_index[i+1]=y
#                             i=i+2
                                   
                        
                    if not flag_create:
                        w_tem = AddW(wgb_locals[0], wgb_locals[1])
                        w_glob=AvgW(w_tem,2)
                    else:
                        #print(unsup_confuse_matrix_index)
                        w_tem = AddW(wgb_locals[0], wgb_locals[1])
                        w1_extra=copy.deepcopy(w_tem)
                        w1=AvgW(w_tem,2)
                        w_tem = AddW(wgb_locals[indecies[0]], wgb_locals[indecies[1]])
                        w_tem = AddW(w1_extra, w_tem)
                        w2=AvgW(w_tem,4)
                        w_tem = AddW(wgb_locals[indecies[2]], wgb_locals[indecies[3]])
                        w_tem = AddW(w1_extra, w_tem)
                        w3=AvgW(w_tem,4)
                        w_tem = AddW(wgb_locals[indecies[4]], wgb_locals[indecies[5]])
                        w_tem = AddW(w1_extra,w_tem)
                        w4=AvgW(w_tem,4)
                        w_tem = AddW(wgb_locals[indecies[6]], wgb_locals[indecies[7]])
                        w_tem = AddW(w1_extra,w_tem)
                        w5=AvgW(w_tem,4)
                        w_tem = AddW(w2, w3)
                        w_tem = AddW(w_tem, w4)
                        w_tem = AddW(w_tem, w5)
                        w_glob=AvgW(w_tem,4)
                else:
                    w_glob = AvgW(w_locals,w_len)
                    w2= copy.deepcopy(w_glob)
                    w3= copy.deepcopy(w_glob)
                    w4= copy.deepcopy(w_glob)
                    w5= copy.deepcopy(w_glob)
                    
            print("begin fedavg")

                

            net_glob.load_state_dict(w_glob)
  
            loss_avg.append(sum(loss_locals) / len(loss_locals))
            accuracy_avg.append(sum(Accurcy_locals) / len(Accurcy_locals))
            sup_loss_avg.append(sum(sup_loss_locals) / len(sup_loss_locals))
            sup_accuracy_avg.append(sum(sup_Accurcy_locals) / len(sup_Accurcy_locals))
            if flag_create:
                used_lp_ratio_avg.append(sum(used_lp_ratio) / len(used_lp_ratio))
                unsup_loss_avg.append(sum(unsup_loss_locals) / len(unsup_loss_locals))
                unsup_accuracy_avg.append(sum(unsup_Accurcy_locals) / len(unsup_Accurcy_locals))
            else :
                used_lp_ratio_avg.append(0)
                unsup_loss_avg.append(0)
                unsup_accuracy_avg.append(0)

            # print( "Loss Avg {}, Acc Avg {} Round {}".format(loss_avg,Accuracy_avg, com_round))
            if not flag_create:
                logging.info(
                    "Loss Avg {}, Accuracy Avg {}, supLoss {}, supAcc {},  Round {} LR {} ".format(
                        loss_avg[com_round], accuracy_avg[com_round], sup_loss_avg[com_round],
                        sup_accuracy_avg[com_round],  com_round,
                        args.base_lr)
                )
            else:

                logging.info(
                    "Loss Avg {}, Accuracy Avg {}, supLoss {}, supAcc {}, unsupLoss {},  unsupAcc {}, Round {} LR {} ".format(
                        loss_avg[com_round], accuracy_avg[com_round], sup_loss_avg[com_round],
                        sup_accuracy_avg[com_round], unsup_loss_avg[com_round], unsup_accuracy_avg[com_round], com_round,
                        args.base_lr)
                )
            if com_round % 10 == 0:
                save_mode_path = os.path.join(
                    snapshot_folder, "epoch_" + str(com_round) + ".pth"
                )
                torch.save(
                    {
                        "state_dict": net_glob.state_dict(),
                    },
                    save_mode_path,
                )
            AUROC_avg, val_acc,Accus_avg, Senss_avg, Specs_avg = test(com_round, w_glob)
            if(Accus_avg>higher_accuracy and com_round>20):
                higher_accuracy_round=com_round
                higher_accuracy=Accus_avg
                logging.info(
                "\nbest accuracy round: {}".format(
                    com_round
                    )
                )
                save_mode_path = os.path.join(
                    snapshot_folder, "best" + ".pth"
                )
                torch.save(
                    {
                        "state_dict": net_glob.state_dict(),
                    },
                    save_mode_path,
                )
            
            val_accuracy.append(val_acc)
            val_accuracy_thres.append(Accus_avg)
            # logging.info("\nTEST Student: Epoch: {}".format(com_round))
            logging.info(
                "\nTEST AUROC: {:6f}, TEST Accus: {:6f}, TEST Senss: {:6f}, TEST Specs: {:6f}".format(
                    AUROC_avg, Accus_avg, Senss_avg, Specs_avg
                )
            )

        
        
        saveTraningValidationResults(new_folder_path,accuracy_avg,sup_accuracy_avg,unsup_accuracy_avg,val_accuracy,val_accuracy_thres,used_lp_ratio_avg,loss_avg,sup_loss_avg,unsup_loss_avg)
        checkpoint_path = os.path.join("..", "model", f"{args.dataset}_{args.mode}", "best.pth")
        saving_path=os.path.join(new_folder_path, 'test.csv')
        final_test(checkpoint_path,saving_path,args)
        
    except KeyboardInterrupt:

        saveTraningValidationResults(new_folder_path,accuracy_avg,sup_accuracy_avg,unsup_accuracy_avg,val_accuracy,val_accuracy_thres,used_lp_ratio_avg,loss_avg,sup_loss_avg,unsup_loss_avg)